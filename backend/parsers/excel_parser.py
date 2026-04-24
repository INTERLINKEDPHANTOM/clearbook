from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def parse_excel(file_path: Path, preview_rows: int = 20) -> dict[str, Any]:
    if not file_path.exists() or file_path.stat().st_size == 0:
        raise ValueError("File appears to be empty or unreadable")

    try:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("File appears to be empty or unreadable") from exc

    try:
        worksheet = workbook.active
        row_iterator = worksheet.iter_rows(values_only=True)

        header_row = next(row_iterator, None)
        if header_row is None:
            raise ValueError("File appears to be empty or unreadable")

        if all(value is None or not str(value).strip() for value in header_row):
            raise ValueError("File is missing required columns")

        if any(value is None or not str(value).strip() for value in header_row):
            raise ValueError("File is missing required columns")

        headers: list[str] = []
        for index, value in enumerate(header_row, start=1):
            if value is None or not str(value).strip():
                headers.append(f"column_{index}")
            else:
                headers.append(str(value).strip())

        preview: list[dict[str, Any]] = []
        row_count = 0

        for row in row_iterator:
            if row is None or all(cell is None or not str(cell).strip() for cell in row):
                continue

            row_count += 1
            if len(preview) >= preview_rows:
                continue

            normalized_row = list(row)
            if len(normalized_row) < len(headers):
                normalized_row.extend([None] * (len(headers) - len(normalized_row)))

            preview.append(
                {
                    headers[i]: _normalize_value(normalized_row[i])
                    for i in range(len(headers))
                }
            )

        if row_count == 0:
            raise ValueError("File appears to be empty or unreadable")

        return {
            "columns": headers,
            "row_count": row_count,
            "preview": preview,
        }
    finally:
        workbook.close()
